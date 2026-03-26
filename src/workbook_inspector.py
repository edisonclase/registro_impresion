from __future__ import annotations

from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .sheet_filters import (
    classify_sheet,
    find_stop_sheet,
    get_printable_sheet_names,
)


@dataclass
class SheetReport:
    index: int
    name: str
    kind: str
    max_row: int
    max_column: int
    print_area: str | None
    orientation: str | None
    paper_size: Any
    is_hidden: bool
    likely_printable: bool
    notes: list[str]


@dataclass
class WorkbookInspectionResult:
    workbook_name: str
    workbook_path: str
    total_sheets: int
    stop_sheet_configured: str
    detected_stop_sheet: str | None
    printable_sheet_count: int
    printable_sheet_names: list[str]
    excluded_sheet_names: list[str]
    sheet_reports: list[SheetReport]
    summary_by_kind: dict[str, int]


def build_sheet_notes(
    *,
    kind: str,
    max_row: int,
    max_column: int,
    is_hidden: bool,
) -> tuple[bool, list[str]]:
    notes: list[str] = []
    likely_printable = True

    if is_hidden:
        notes.append("hoja oculta")
        likely_printable = False

    if max_row <= 10 and max_column <= 6:
        notes.append("hoja muy pequeña; posible auxiliar")
        likely_printable = False

    if kind == "calificaciones":
        notes.append("hoja candidata para impresión horizontal de lado a lado")
        notes.append("en fase posterior: excluir nombres y conservar número, encabezados y notas")

    if kind == "datos_estudiante":
        notes.append("hoja de formulario con texto oficial; no modificar contenido")

    if kind == "datos_centro":
        notes.append("hoja institucional; no modificar contenido")

    if kind == "asistencia":
        notes.append("revisar ajuste de ancho/alto para mantener espacios vacíos visibles")

    if kind in {"completivo", "extraordinario", "acta"}:
        notes.append("revisar contra PDF físico del grado antes de ajustar impresión")

    return likely_printable, notes


def inspect_workbook(workbook_path: Path, stop_sheet_name: str) -> WorkbookInspectionResult:
    wb = load_workbook(workbook_path, data_only=False)

    all_sheet_names = wb.sheetnames
    detected_stop_sheet = find_stop_sheet(all_sheet_names, stop_sheet_name)
    printable_sheet_names = get_printable_sheet_names(all_sheet_names, stop_sheet_name)
    excluded_sheet_names = all_sheet_names[len(printable_sheet_names):]

    sheet_reports: list[SheetReport] = []
    summary_by_kind: dict[str, int] = {}

    for idx, sheet_name in enumerate(printable_sheet_names, start=1):
        ws = wb[sheet_name]

        print_area = None
        try:
            if ws.print_area:
                print_area = str(ws.print_area)
        except Exception:
            print_area = None

        orientation = None
        try:
            orientation = ws.page_setup.orientation
        except Exception:
            orientation = None

        paper_size = None
        try:
            paper_size = ws.page_setup.paperSize
        except Exception:
            paper_size = None

        is_hidden = ws.sheet_state != "visible"
        kind = classify_sheet(sheet_name)

        likely_printable, notes = build_sheet_notes(
            kind=kind,
            max_row=ws.max_row,
            max_column=ws.max_column,
            is_hidden=is_hidden,
        )

        summary_by_kind[kind] = summary_by_kind.get(kind, 0) + 1

        sheet_reports.append(
            SheetReport(
                index=idx,
                name=sheet_name,
                kind=kind,
                max_row=ws.max_row,
                max_column=ws.max_column,
                print_area=print_area,
                orientation=orientation,
                paper_size=paper_size,
                is_hidden=is_hidden,
                likely_printable=likely_printable,
                notes=notes,
            )
        )

    return WorkbookInspectionResult(
        workbook_name=workbook_path.name,
        workbook_path=str(workbook_path),
        total_sheets=len(all_sheet_names),
        stop_sheet_configured=stop_sheet_name,
        detected_stop_sheet=detected_stop_sheet,
        printable_sheet_count=len(printable_sheet_names),
        printable_sheet_names=printable_sheet_names,
        excluded_sheet_names=excluded_sheet_names,
        sheet_reports=sheet_reports,
        summary_by_kind=summary_by_kind,
    )


def inspection_result_to_dict(result: WorkbookInspectionResult) -> dict[str, Any]:
    return {
        "workbook_name": result.workbook_name,
        "workbook_path": result.workbook_path,
        "total_sheets": result.total_sheets,
        "stop_sheet_configured": result.stop_sheet_configured,
        "detected_stop_sheet": result.detected_stop_sheet,
        "printable_sheet_count": result.printable_sheet_count,
        "printable_sheet_names": result.printable_sheet_names,
        "excluded_sheet_names": result.excluded_sheet_names,
        "summary_by_kind": result.summary_by_kind,
        "sheet_reports": [asdict(item) for item in result.sheet_reports],
    }