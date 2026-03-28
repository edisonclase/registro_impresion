from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetPrintAction:
    sheet_name: str
    detected_kind: str
    action_taken: list[str]


THIN_BLACK_SIDE = Side(style="thin", color="FF000000")
THIN_BLACK_BORDER = Border(
    left=THIN_BLACK_SIDE,
    right=THIN_BLACK_SIDE,
    top=THIN_BLACK_SIDE,
    bottom=THIN_BLACK_SIDE,
)

WHITE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFFFFF",
    bgColor="FFFFFFFF",
)


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def clear_conditional_formatting(ws: Worksheet) -> list[str]:
    removed = 0

    try:
        removed = len(ws.conditional_formatting)
    except Exception:
        removed = 0

    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass

    if removed:
        return [f"formato condicional eliminado: {removed} reglas"]
    return ["sin formato condicional que eliminar"]


def font_to_times_new_roman_black(original_font: Font) -> Font:
    font_copy = copy(original_font)
    font_copy.name = "Times New Roman"
    font_copy.size = 12
    font_copy.color = "FF000000"
    return font_copy


def white_fill() -> PatternFill:
    return copy(WHITE_FILL)


def normalize_cell_visual_style(cell) -> tuple[bool, bool]:
    font_changed = False
    fill_changed = False

    try:
        cell.font = font_to_times_new_roman_black(cell.font)
        font_changed = True
    except Exception:
        pass

    try:
        fill = cell.fill
        fill_type = getattr(fill, "fill_type", None)
        fg_rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
        bg_rgb = getattr(getattr(fill, "bgColor", None), "rgb", None)

        has_colored_fill = (
            fill_type is not None
            or fg_rgb not in (None, "00000000", "00FFFFFF", "FFFFFFFF")
            or bg_rgb not in (None, "00000000", "00FFFFFF", "FFFFFFFF")
        )

        if has_colored_fill:
            cell.fill = white_fill()
            fill_changed = True
    except Exception:
        pass

    return font_changed, fill_changed


def normalize_sheet_visual_style(ws: Worksheet) -> list[str]:
    changed_fill_count = 0
    changed_font_count = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell is None:
                continue

            font_changed, fill_changed = normalize_cell_visual_style(cell)

            if font_changed:
                changed_font_count += 1
            if fill_changed:
                changed_fill_count += 1

    notes = ["fuente global forzada a Times New Roman 12 y color negro"]
    if changed_fill_count:
        notes.append(f"rellenos forzados a blanco: {changed_fill_count}")
    if changed_font_count:
        notes.append(f"fuentes forzadas a negro: {changed_font_count}")
    return notes


def is_row_hidden(ws: Worksheet, row_idx: int) -> bool:
    return bool(ws.row_dimensions[row_idx].hidden)


def is_col_hidden(ws: Worksheet, col_idx: int) -> bool:
    col_letter = get_column_letter(col_idx)
    return bool(ws.column_dimensions[col_letter].hidden)


def get_used_range_visible_bounds(ws: Worksheet) -> tuple[int, int]:
    """
    Busca el último renglón y columna con contenido real.
    Ayuda a no arrastrar columnas auxiliares vacías al print_area.
    """
    last_row = 1
    last_col = 1

    for row_idx in range(1, ws.max_row + 1):
        row_has_data = False
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                row_has_data = True
                last_col = max(last_col, col_idx)
        if row_has_data:
            last_row = row_idx

    return max(1, last_row), max(1, last_col)


def set_print_area_used_range(ws: Worksheet) -> str:
    last_row, last_col = get_used_range_visible_bounds(ws)
    last_col_letter = get_column_letter(last_col)
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    return f"A1:{last_col_letter}{last_row}"


def autosize_columns_by_content(
    ws: Worksheet,
    *,
    min_width: float = 8.5,
    max_width: float = 45.0,
) -> list[str]:
    """
    Ajusta solo columnas visibles.
    No modifica columnas ocultas.
    """
    changed = 0

    for col_idx in range(1, ws.max_column + 1):
        if is_col_hidden(ws, col_idx):
            continue

        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue

            text = str(value).replace("\n", " ").strip()
            if not text:
                continue

            max_len = max(max_len, len(text))

        if max_len == 0:
            continue

        proposed = min(max(min_width, max_len * 0.95), max_width)

        current_width = ws.column_dimensions[col_letter].width
        if current_width is None or proposed > current_width:
            ws.column_dimensions[col_letter].width = proposed
            changed += 1

    return [f"ancho de columnas visibles ajustado según contenido: {changed} columnas"]


def estimate_text_lines(text: str, width_units: float) -> int:
    if not text:
        return 1

    width_units = max(width_units, 6.0)
    approx_chars_per_line = max(6, int(width_units * 1.15))
    return max(1, (len(text) // approx_chars_per_line) + (1 if len(text) % approx_chars_per_line else 0))


def preserve_and_adjust_row_heights(
    ws: Worksheet,
    *,
    min_height: float = 18.0,
    wrapped_min_height: float = 24.0,
    rotated_min_height: float = 42.0,
    max_height: float = 96.0,
    max_scan_header_cols: int | None = None,
) -> list[str]:
    """
    Conserva alturas existentes y las aumenta solo cuando hace falta.
    Protege texto largo, wrap_text y texto vertical/rotado.
    No toca filas ocultas.
    """
    changed_rows = 0
    rotated_rows = 0
    wrapped_rows = 0

    last_row, last_col = get_used_range_visible_bounds(ws)
    scan_cols = min(last_col, max_scan_header_cols) if max_scan_header_cols else last_col

    for row_idx in range(1, last_row + 1):
        if is_row_hidden(ws, row_idx):
            continue

        row_dim = ws.row_dimensions[row_idx]
        current_height = row_dim.height if row_dim.height is not None else 15.0
        target_height = max(current_height, min_height)

        found_wrapped = False
        found_rotated = False

        for col_idx in range(1, scan_cols + 1):
            if is_col_hidden(ws, col_idx):
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value in (None, ""):
                continue

            text = str(value).strip()
            if not text:
                continue

            alignment = cell.alignment
            rotation = getattr(alignment, "textRotation", 0) or 0
            wrap_text = bool(getattr(alignment, "wrapText", False))

            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            if width is None:
                width = 8.43

            if rotation not in (0, None):
                found_rotated = True
                target_height = max(target_height, rotated_min_height)

                if len(text) > 6:
                    extra = min(max_height, rotated_min_height + (len(text) * 1.2))
                    target_height = max(target_height, extra)

            if wrap_text or len(text) > max(12, int(width * 1.3)):
                found_wrapped = True
                lines = estimate_text_lines(text, width)
                estimated_height = min(max_height, wrapped_min_height + ((lines - 1) * 12))
                target_height = max(target_height, estimated_height)

        target_height = min(max_height, target_height)

        if target_height > current_height + 0.1:
            row_dim.height = target_height
            changed_rows += 1

        if found_rotated:
            rotated_rows += 1
        if found_wrapped:
            wrapped_rows += 1

    notes = [f"alturas de fila ajustadas conservando valores originales: {changed_rows} filas"]
    if rotated_rows:
        notes.append(f"filas con texto vertical/rotado protegidas: {rotated_rows}")
    if wrapped_rows:
        notes.append(f"filas con texto largo o envuelto protegidas: {wrapped_rows}")
    return notes


def complete_used_range_borders(ws: Worksheet) -> list[str]:
    applied = 0
    last_row, last_col = get_used_range_visible_bounds(ws)

    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            if cell is None:
                continue
            try:
                cell.border = copy(THIN_BLACK_BORDER)
                applied += 1
            except Exception:
                pass

    return [f"bordes completos aplicados en rango usado: {applied} celdas"]

def find_header_columns(
    ws: Worksheet,
    header_candidates: set[str],
    max_header_row: int = 8,
) -> list[int]:
    """
    Busca columnas por encabezado en las primeras filas.
    Devuelve índices de columna.
    """
    found: list[int] = []
    normalized_targets = {h.replace(" ", "") for h in header_candidates}

    for col_idx in range(1, ws.max_column + 1):
        for row_idx in range(1, min(ws.max_row, max_header_row) + 1):
            text = cell_text(ws.cell(row=row_idx, column=col_idx).value)
            compact = text.replace(" ", "")
            if text in header_candidates or compact in normalized_targets:
                found.append(col_idx)
                break

    return found


def looks_like_subject_competency_sheet(ws: Worksheet, max_header_row: int = 8) -> bool:
    """
    Detecta hojas tipo asignatura/competencia por estructura,
    no por el nombre de la pestaña.

    Patrón esperado:
    - existe columna NOMBRE / NOMBRES
    - existen encabezados de períodos como P1/P2/P3/P4 o RP1/RP2...
    """
    name_headers = {"NOMBRE", "NOMBRES", "NOMBRE DEL ESTUDIANTE"}
    period_headers = {"P1", "P2", "P3", "P4", "RP1", "RP2", "RP3", "RP4"}

    name_cols = find_header_columns(ws, name_headers, max_header_row=max_header_row)
    period_cols = find_header_columns(ws, period_headers, max_header_row=max_header_row)

    # Para evitar falsos positivos:
    # - debe existir nombre
    # - deben existir varias columnas de período
    # - la hoja debe tener tamaño razonable
    return bool(name_cols) and len(period_cols) >= 4 and ws.max_row >= 10 and ws.max_column >= 8


def hide_name_column_for_subject_sheet(ws: Worksheet, max_header_row: int = 8) -> list[str]:
    """
    Oculta la columna del nombre por encabezado.
    Si no encuentra encabezado, usa B como fallback para no romper lo ya funcional.
    """
    actions: list[str] = []
    name_headers = {"NOMBRE", "NOMBRES", "NOMBRE DEL ESTUDIANTE"}

    name_cols = find_header_columns(ws, name_headers, max_header_row=max_header_row)

    if name_cols:
        hidden_letters: list[str] = []
        for col_idx in name_cols:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True
            hidden_letters.append(col_letter)
        actions.append(f"columna(s) de nombre ocultas por encabezado: {', '.join(hidden_letters)}")
        return actions

    ws.column_dimensions["B"].hidden = True
    actions.append("columna B oculta como fallback para no imprimir nombres")
    return actions


def force_wrap_text_used_range(ws: Worksheet) -> list[str]:
    """
    Fuerza wrap_text dentro del rango usado.
    Útil para CEILE y hojas con mucho texto.
    """
    changed = 0
    last_row, last_col = get_used_range_visible_bounds(ws)

    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            if cell is None:
                continue
            if cell.value in (None, ""):
                continue

            try:
                alignment = copy(cell.alignment)
                alignment.wrapText = True
                cell.alignment = alignment
                changed += 1
            except Exception:
                pass

    return [f"wrap_text forzado en rango usado: {changed} celdas"]

def looks_like_subject_score_layout(ws: Worksheet, max_header_row: int = 8) -> bool:
    """
    Detecta hojas tipo calificaciones por asignatura aunque no entren
    por la detección anterior de competencias.

    Señales:
    - existe encabezado NOMBRE/NOMBRES
    - existen varios encabezados de periodos P1/P2/P3/P4 o RP1/RP2...
    """
    name_headers = {"NOMBRE", "NOMBRES", "NOMBRE DEL ESTUDIANTE"}
    period_headers = {"P1", "P2", "P3", "P4", "RP1", "RP2", "RP3", "RP4"}

    name_cols = find_header_columns(ws, name_headers, max_header_row=max_header_row)
    period_cols = find_header_columns(ws, period_headers, max_header_row=max_header_row)

    return bool(name_cols) and len(period_cols) >= 4 and ws.max_row >= 10

def hide_student_name_columns_by_header(ws: Worksheet, max_header_row: int = 8) -> list[str]:
    """
    Oculta columnas cuyo encabezado diga NOMBRE / NOMBRES.
    No asume una letra fija; busca por encabezado.
    """
    hidden_letters: list[str] = []

    target_headers = {"NOMBRE", "NOMBRES", "NOMBRE DEL ESTUDIANTE", "ESTUDIANTE"}

    for col_idx in range(1, ws.max_column + 1):
        for row_idx in range(1, min(ws.max_row, max_header_row) + 1):
            text = cell_text(ws.cell(row=row_idx, column=col_idx).value)
            if text in target_headers:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].hidden = True
                hidden_letters.append(col_letter)
                break

    if hidden_letters:
        return [f"columnas de nombre ocultas por encabezado: {', '.join(hidden_letters)}"]

    return ["no se encontraron columnas de nombre por encabezado"]

def apply_subject_name_hiding_fallback(ws: Worksheet) -> list[str]:
    """
    Corrección transversal:
    si una hoja tiene estructura de calificaciones por asignatura,
    ocultar la columna del nombre aunque no haya sido clasificada
    explícitamente como 'competencias'.
    """
    if looks_like_subject_score_layout(ws):
        return hide_name_column_for_subject_sheet(ws)

    return ["sin corrección transversal de nombre por estructura"]

def clamp_visible_column_widths(
    ws: Worksheet,
    *,
    min_width: float = 6.5,
    max_width: float = 18.0,
) -> list[str]:
    """
    Limita el ancho de columnas visibles para que el PDF no se deforme.
    Solo reduce columnas que se hayan ido demasiado anchas.
    """
    changed = 0

    for col_idx in range(1, ws.max_column + 1):
        if is_col_hidden(ws, col_idx):
            continue

        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width

        if width is None:
            continue

        new_width = max(min_width, min(width, max_width))
        if abs(new_width - width) > 0.1:
            ws.column_dimensions[col_letter].width = new_width
            changed += 1

    return [f"ancho de columnas visibles limitado para PDF: {changed} columnas"]

def configure_ce_sheet_for_print(ws: Worksheet) -> list[str]:
    """
    Hojas que comienzan con CE:
    - ocultar nombres
    - mantener impresión vertical
    - ajuste conservador para PDF
    - completar bordes del rango usado
    """
    actions: list[str] = []

    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")
    actions.append("ajuste a 1 página de ancho")

    actions.extend(hide_student_name_columns_by_header(ws))

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    ws.print_title_rows = "$1:$6"
    actions.append("filas 1 a 6 repetidas como encabezado")

    actions.extend(complete_used_range_borders(ws))
    actions.extend(autosize_columns_by_content(ws, min_width=6.5, max_width=16.0))
    actions.extend(clamp_visible_column_widths(ws, min_width=6.5, max_width=16.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=18.0,
            wrapped_min_height=24.0,
            rotated_min_height=42.0,
            max_height=72.0,
        )
    )
    return actions

def configure_ceile_sheet_for_print(ws: Worksheet) -> list[str]:
    """
    Hojas CEILE:
    - ocultar nombres
    - ajustar ancho y alto al contenido con límites finos
    - completar bordes
    - forzar wrap_text para evitar cortes
    """
    actions: list[str] = []

    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")
    actions.append("ajuste a 1 página de ancho")

    actions.extend(hide_student_name_columns_by_header(ws))

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    ws.print_title_rows = "$1:$6"
    actions.append("filas 1 a 6 repetidas como encabezado")

    actions.extend(force_wrap_text_used_range(ws))
    actions.extend(complete_used_range_borders(ws))

    # más conservador en ancho, pero menos agresivo al cerrar
    actions.extend(autosize_columns_by_content(ws, min_width=7.0, max_width=16.0))
    actions.extend(clamp_visible_column_widths(ws, min_width=7.0, max_width=16.0))

    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=22.0,
            wrapped_min_height=30.0,
            rotated_min_height=50.0,
            max_height=110.0,
        )
    )
    return actions


def detect_sheet_kind_by_content(ws: Worksheet) -> str:
    title = cell_text(ws.title)

    a1 = cell_text(ws["A1"].value)
    b1 = cell_text(ws["B1"].value)
    d1 = cell_text(ws["D1"].value)

    a2 = cell_text(ws["A2"].value)
    b2 = cell_text(ws["B2"].value)

    d4 = cell_text(ws["D4"].value)
    c6 = cell_text(ws["C6"].value)

    row4 = " | ".join(cell_text(ws.cell(4, c).value) for c in range(1, min(ws.max_column, 12) + 1))

    if "DATOS DEL CENTRO" in b1:
        return "datos_centro"

    if "DATOS DEL ESTUDIANTE" in b1:
        return "datos_estudiante"

    if title.startswith("ECAP"):
        return "ecap"

    if title.startswith("CF-") or "C.F." in row4:
        return "cf"

    if title.startswith("CEILE"):
        return "ceile"

    if title.startswith("CE"):
        return "ce"

    # NUEVO: detectar hojas de asignaturas por estructura,
    # aunque no se llamen LE, MAT, NAT, etc.
    if looks_like_subject_competency_sheet(ws):
        return "competencias"

    # Se deja esta regla vieja por compatibilidad
    if "COMPETENCIA" in a2 and "NOMBRE" in b2:
        return "competencias"

    if "DÍAS TRABAJADOS" in d4 and "NOMBRE" in c6:
        return "asistencia_asignatura"

    if "COMPLETIVO" in a1 or "COMPLETIVO" in b1 or "COMPLETIVO" in d1:
        return "completivo"

    if "EXTRAORDINARIO" in a1 or "EXTRAORDINARIO" in b1 or "EXTRAORDINARIO" in d1:
        return "extraordinario"

    return "general"


def set_common_page_setup_letter_portrait(ws: Worksheet) -> None:
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.20
    ws.page_margins.right = 0.20
    ws.page_margins.top = 0.30
    ws.page_margins.bottom = 0.30
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False


def find_cf_attendance_columns(ws: Worksheet, max_header_row: int = 8) -> tuple[list[int], list[int]]:
    attendance_cols: list[int] = []
    annual_cols: list[int] = []

    attendance_markers = {
        "P1",
        "P2",
        "P3",
        "P4",
        "% ANUAL",
        "%ANUAL",
        "ASISTENCIA",
    }

    annual_markers = {
        "% ANUAL",
        "%ANUAL",
    }

    normalized_attendance = {m.replace(" ", "") for m in attendance_markers}
    normalized_annual = {m.replace(" ", "") for m in annual_markers}

    for col_idx in range(1, ws.max_column + 1):
        matched_texts: set[str] = set()

        for row_idx in range(1, min(ws.max_row, max_header_row) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            text = cell_text(value)
            compact = text.replace(" ", "")
            if text:
                matched_texts.add(text)
                matched_texts.add(compact)

        if matched_texts.intersection(attendance_markers.union(normalized_attendance)):
            attendance_cols.append(col_idx)

        if matched_texts.intersection(annual_markers.union(normalized_annual)):
            annual_cols.append(col_idx)

    return attendance_cols, annual_cols


def hide_cf_attendance_details_except_annual(ws: Worksheet) -> list[str]:
    """
    Regla fija para hojas CF del formato actual:
    en el bloque ASISTENCIA se dejan ocultas J:M
    y se deja visible N (% ANUAL).
    """
    actions: list[str] = []

    cols_to_hide = ["J", "K", "L", "M"]
    hidden_count = 0

    for col_letter in cols_to_hide:
        if ws.max_column >= ws[col_letter + "1"].column:
            ws.column_dimensions[col_letter].hidden = True
            hidden_count += 1

    if ws.max_column >= ws["N1"].column:
        ws.column_dimensions["N"].hidden = False
        actions.append("columna N (% ANUAL) forzada visible")

    actions.append(f"bloque de asistencia CF ajustado manualmente: {hidden_count} columnas ocultas (J:M)")
    return actions


def configure_competency_sheet_for_print(ws: Worksheet) -> list[str]:
    actions: list[str] = []

    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")
    actions.append("ajuste a 1 página de ancho")

    # NUEVO: ocultar nombre por encabezado, no por letra fija únicamente
    actions.extend(hide_name_column_for_subject_sheet(ws))

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    ws.print_title_rows = "$1:$4"
    actions.append("filas 1 a 4 repetidas como encabezado")

    actions.extend(autosize_columns_by_content(ws, min_width=8.5, max_width=30.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=18.0,
            wrapped_min_height=24.0,
            rotated_min_height=42.0,
            max_height=96.0,
        )
    )
    return actions


def configure_cf_sheet_for_print(ws: Worksheet) -> list[str]:
    actions: list[str] = []

    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")
    actions.append("ajuste a 1 página de ancho")

    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["D"].hidden = True
    actions.append("columnas B y D ocultas para no imprimir ID ni nombre")

    actions.extend(hide_cf_attendance_details_except_annual(ws))

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    ws.print_title_rows = "$1:$6"
    actions.append("filas 1 a 6 repetidas como encabezado")

    actions.extend(autosize_columns_by_content(ws, min_width=8.5, max_width=24.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=18.0,
            wrapped_min_height=24.0,
            rotated_min_height=44.0,
            max_height=110.0,
        )
    )
    return actions


def configure_attendance_sheet_for_print(ws: Worksheet) -> list[str]:
    actions: list[str] = []

    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")
    actions.append("ajuste a 1 página de ancho")

    for col in ["A", "C", "AA", "AB", "AC", "AD"]:
        ws.column_dimensions[col].hidden = True

    for col in ["AF", "BD", "BE", "BF", "BG"]:
        ws.column_dimensions[col].hidden = True

    actions.append("columnas ocultas en asistencia: ID, nombre, TA, %A, TE, %E")

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    ws.print_title_rows = "$1:$6"
    actions.append("filas 1 a 6 repetidas como encabezado")

    actions.extend(autosize_columns_by_content(ws, min_width=4.5, max_width=18.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=18.0,
            wrapped_min_height=24.0,
            rotated_min_height=44.0,
            max_height=110.0,
        )
    )
    return actions


def configure_text_sheet_for_print(ws: Worksheet) -> list[str]:
    actions: list[str] = []
    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    actions.extend(autosize_columns_by_content(ws, min_width=8.5, max_width=28.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=20.0,
            wrapped_min_height=28.0,
            rotated_min_height=48.0,
            max_height=120.0,
        )
    )
    return actions


def configure_ecap_sheet_for_print(ws: Worksheet) -> list[str]:
    actions: list[str] = []
    set_common_page_setup_letter_portrait(ws)
    actions.append("orientación vertical en carta")

    print_area = set_print_area_used_range(ws)
    actions.append(f"área de impresión definida: {print_area}")

    actions.extend(complete_used_range_borders(ws))
    actions.extend(autosize_columns_by_content(ws, min_width=12.0, max_width=42.0))
    actions.extend(
        preserve_and_adjust_row_heights(
            ws,
            min_height=18.0,
            wrapped_min_height=24.0,
            rotated_min_height=44.0,
            max_height=96.0,
        )
    )
    return actions


def prepare_print_workbook(
    source_path: Path,
    output_path: Path,
    stop_sheet_name: Optional[str] = None,
) -> list[SheetPrintAction]:
    wb = load_workbook(source_path)

    actions_log: list[SheetPrintAction] = []
    stop_detected = False

    stop_tokens = ["REPORTE CALIFICACI"]
    if stop_sheet_name:
        stop_tokens.append(stop_sheet_name.upper())

    for ws in wb.worksheets:
        title_upper = (ws.title or "").upper()

        if any(token in title_upper for token in stop_tokens):
            stop_detected = True

        if stop_detected:
            actions_log.append(
                SheetPrintAction(
                    sheet_name=ws.title,
                    detected_kind="excluida_post_corte",
                    action_taken=["sin cambios en esta fase"],
                )
            )
            continue

        kind = detect_sheet_kind_by_content(ws)
        actions: list[str] = []

        actions.extend(clear_conditional_formatting(ws))
        actions.extend(normalize_sheet_visual_style(ws))
        actions.extend(apply_subject_name_hiding_fallback(ws))

        if kind == "competencias":
            actions.extend(configure_competency_sheet_for_print(ws))
        elif kind == "cf":
            actions.extend(configure_cf_sheet_for_print(ws))
        elif kind == "ecap":
            actions.extend(configure_ecap_sheet_for_print(ws))
        elif kind == "ceile":
            actions.extend(configure_ceile_sheet_for_print(ws))
        elif kind == "ce":
            actions.extend(configure_ce_sheet_for_print(ws))
        elif kind == "asistencia_asignatura":
            actions.extend(configure_attendance_sheet_for_print(ws))
        else:
            actions.extend(configure_text_sheet_for_print(ws))

        actions_log.append(
            SheetPrintAction(
                sheet_name=ws.title,
                detected_kind=kind,
                action_taken=actions,
            )
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return actions_log